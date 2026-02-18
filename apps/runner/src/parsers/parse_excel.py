"""
Excel 入力パーサ — 横カレンダーシートからランクマトリクスを抽出する。

Usage:
    python parse_excel.py <excel_path> [--sheet <sheet_name>]

Output:
    JSON to stdout:
    {
      "facility_name_candidate": "プールヴィラ古宇利島",
      "sheet_name": "横カレンダー",
      "dates": ["2025-11-01", ...],
      "room_types": ["ラグジュアリープールヴィラ(単泊)", ...],
      "ranks": [
        {"date": "2025-11-01", "room_type": "...", "rank_code": "H"},
        ...
      ],
      "warnings": ["C4 is blank — no rank for ..."],
      "meta": {
        "total_cells": 848,
        "filled_cells": 826,
        "empty_cells": 22,
        "right_boundary": "KT",
        "down_boundary": 7
      }
    }

Reference: docs/design.md §4
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 定数
# ---------------------------------------------------------------------------
DEFAULT_SHEET_NAME = "横カレンダー"
DATE_ROW = 3          # 日付が並ぶ行
ROOM_TYPE_COL = 2     # 部屋タイプ名の列 (B)
DATA_START_ROW = 4    # データ開始行
DATA_START_COL = 3    # データ開始列 (C)

FACILITY_NAME_RE = re.compile(r"【(.+?)様】")


# ---------------------------------------------------------------------------
# ヘルパー
# ---------------------------------------------------------------------------
def extract_facility_name(filename: str) -> str | None:
    """ファイル名から施設名候補を抽出する。"""
    m = FACILITY_NAME_RE.search(filename)
    return m.group(1) if m else None


def normalize_rank(value: object) -> str | None:
    """
    セル値をランクコード文字列に正規化する。
    - None / 空文字 → None (空セル)
    - 文字列 → trim + 全角→半角 + 大文字化
    """
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    # 全角英字 → 半角英字
    result = ""
    for ch in s:
        cp = ord(ch)
        if 0xFF21 <= cp <= 0xFF3A:       # Ａ-Ｚ
            result += chr(cp - 0xFF21 + ord("A"))
        elif 0xFF41 <= cp <= 0xFF5A:     # ａ-ｚ
            result += chr(cp - 0xFF41 + ord("A"))
        elif 0xFF10 <= cp <= 0xFF19:     # ０-９
            result += chr(cp - 0xFF10 + ord("0"))
        else:
            result += ch
    return result.upper()


def format_date(value: object) -> str | None:
    """
    セル値を YYYY-MM-DD 文字列に変換する。
    openpyxl は日付セルを datetime オブジェクトとして返す。
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return None
    # "2025/4/1" or "2025-4-1" 形式
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def is_valid_room_type(value: object) -> bool:
    """部屋タイプ名として有効かどうか判定する。"""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return False  # 0 やプレースホルダ数値を除外
    s = str(value).strip()
    return len(s) > 0


# ---------------------------------------------------------------------------
# メインパーサ
# ---------------------------------------------------------------------------
def parse_excel(
    filepath: str,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> dict:
    """Excel ファイルを解析し、ランクマトリクス JSON を返す。"""
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    wb = openpyxl.load_workbook(str(path), data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f'Sheet "{sheet_name}" not found. '
            f"Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    warnings: list[str] = []

    # --- 施設名候補 ---
    facility_name = extract_facility_name(path.name)

    # --- 右端検出 (row 3 の日付ヘッダーをスキャン) ---
    dates: list[str] = []
    col = DATA_START_COL
    while True:
        raw = ws.cell(row=DATE_ROW, column=col).value
        d = format_date(raw)
        if d is None:
            break
        dates.append(d)
        col += 1
    right_col = col - 1  # 最後の有効列

    if not dates:
        raise ValueError(
            f"No dates found in row {DATE_ROW} starting from "
            f"{get_column_letter(DATA_START_COL)}{DATE_ROW}"
        )

    # --- 下端検出 (column B の部屋タイプをスキャン) ---
    room_types: list[str] = []
    row = DATA_START_ROW
    while True:
        raw = ws.cell(row=row, column=ROOM_TYPE_COL).value
        if not is_valid_room_type(raw):
            break
        room_types.append(str(raw).strip())
        row += 1
    bottom_row = row - 1

    if not room_types:
        raise ValueError(
            f"No room types found in column "
            f"{get_column_letter(ROOM_TYPE_COL)} starting from row {DATA_START_ROW}"
        )

    # --- ランクデータ抽出 ---
    ranks: list[dict[str, str]] = []
    filled = 0
    empty = 0

    for ri, room_type in enumerate(room_types):
        data_row = DATA_START_ROW + ri
        for ci, date in enumerate(dates):
            data_col = DATA_START_COL + ci
            raw = ws.cell(row=data_row, column=data_col).value
            rank = normalize_rank(raw)

            if rank is None:
                empty += 1
                cell_ref = f"{get_column_letter(data_col)}{data_row}"
                warnings.append(
                    f"{cell_ref} is blank — no rank for "
                    f"{room_type} on {date}"
                )
            else:
                filled += 1
                ranks.append({
                    "date": date,
                    "room_type": room_type,
                    "rank_code": rank,
                })

    total = filled + empty

    return {
        "facility_name_candidate": facility_name,
        "sheet_name": sheet_name,
        "dates": dates,
        "room_types": room_types,
        "ranks": ranks,
        "warnings": warnings,
        "meta": {
            "total_cells": total,
            "filled_cells": filled,
            "empty_cells": empty,
            "right_boundary": get_column_letter(right_col),
            "down_boundary": bottom_row,
        },
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main() -> None:
    import warnings as _warnings

    _warnings.filterwarnings("ignore", module="openpyxl")

    parser = argparse.ArgumentParser(
        description="Parse Excel rank matrix from 横カレンダー sheet"
    )
    parser.add_argument("filepath", help="Path to .xlsx file")
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET_NAME,
        help=f"Sheet name (default: {DEFAULT_SHEET_NAME})",
    )
    args = parser.parse_args()

    # Ensure UTF-8 output on Windows
    sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]

    try:
        result = parse_excel(args.filepath, args.sheet)
        json.dump(result, sys.stdout, ensure_ascii=False, indent=2)
        sys.stdout.write("\n")
    except Exception as e:
        error = {"error": str(e), "type": type(e).__name__}
        json.dump(error, sys.stderr, ensure_ascii=False)
        sys.stderr.write("\n")
        sys.exit(1)


if __name__ == "__main__":
    main()
