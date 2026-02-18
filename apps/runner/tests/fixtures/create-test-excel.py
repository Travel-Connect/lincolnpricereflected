"""
Create test Excel fixtures for parser tests.
Usage: python create-test-excel.py <output_dir>
"""

import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl


def create_basic(outdir: Path) -> None:
    """Basic 3 dates × 2 room types with all ranks filled."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "横カレンダー"

    # Row 3: dates in C3, D3, E3
    base = datetime(2025, 12, 1)
    ws["C3"] = base
    ws["D3"] = base + timedelta(days=1)
    ws["E3"] = base + timedelta(days=2)

    # Column B: room types
    ws["B4"] = "スタンダード(単泊)"
    ws["B5"] = "デラックス(連泊)"

    # Rank data: C4-E5
    ws["C4"] = "H"
    ws["D4"] = "G"
    ws["E4"] = "P"
    ws["C5"] = "I"
    ws["D5"] = "K"
    ws["E5"] = "E"

    wb.save(str(outdir / "【テスト施設様】料金変動案_20251201.xlsx"))


def create_with_blanks(outdir: Path) -> None:
    """3 dates × 2 room types with some blank rank cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "横カレンダー"

    base = datetime(2025, 12, 1)
    ws["C3"] = base
    ws["D3"] = base + timedelta(days=1)
    ws["E3"] = base + timedelta(days=2)

    ws["B4"] = "Room A"
    ws["B5"] = "Room B"

    # Some blanks
    ws["C4"] = "H"
    # D4 is blank (None)
    ws["E4"] = "G"
    ws["C5"] = None  # explicit blank
    ws["D5"] = "K"
    ws["E5"] = ""  # empty string

    wb.save(str(outdir / "【ブランク施設様】料金変動案_20251201.xlsx"))


def create_boundary_test(outdir: Path) -> None:
    """
    Test boundary detection:
    - 5 dates, then blank column (right boundary)
    - 3 room types, then a 0-value row, then blank (down boundary)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "横カレンダー"

    base = datetime(2026, 1, 1)
    for i in range(5):
        ws.cell(row=3, column=3 + i).value = base + timedelta(days=i)
    # Column 8 (H3) is blank — right boundary at column 7 (G)

    ws["B4"] = "タイプ1"
    ws["B5"] = "タイプ2"
    ws["B6"] = "タイプ3"
    ws["B7"] = 0  # placeholder, not a valid room type
    ws["B8"] = None  # blank

    for row in range(4, 7):
        for col in range(3, 8):
            ws.cell(row=row, column=col).value = chr(
                ord("A") + (row - 4) * 5 + (col - 3)
            )

    wb.save(str(outdir / "【境界テスト様】料金変動案_20260101.xlsx"))


def create_empty(outdir: Path) -> None:
    """Sheet exists but no data at C4 origin."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "横カレンダー"
    # Nothing at C3, B4, etc.
    wb.save(str(outdir / "【空施設様】料金変動案_20260101.xlsx"))


def create_no_sheet(outdir: Path) -> None:
    """No 横カレンダー sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(str(outdir / "【シートなし様】料金変動案_20260101.xlsx"))


def create_fullwidth_ranks(outdir: Path) -> None:
    """Full-width and lowercase rank codes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "横カレンダー"

    ws["C3"] = datetime(2026, 1, 1)
    ws["D3"] = datetime(2026, 1, 2)

    ws["B4"] = "テスト部屋"

    ws["C4"] = "Ａ"   # full-width A
    ws["D4"] = "ｂ"   # full-width lowercase b

    wb.save(str(outdir / "【全角テスト様】料金変動案_20260101.xlsx"))


if __name__ == "__main__":
    outdir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(".")
    outdir.mkdir(parents=True, exist_ok=True)

    create_basic(outdir)
    create_with_blanks(outdir)
    create_boundary_test(outdir)
    create_empty(outdir)
    create_no_sheet(outdir)
    create_fullwidth_ranks(outdir)

    print(f"Created 6 test fixtures in {outdir}")
